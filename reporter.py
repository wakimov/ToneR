# -*- coding: utf-8 -*-
#'''
#Created on June 02 2015 �.
#
#@author: v.akimov
#'''
import openpyxl, datetime 

if __name__ == '__main__':
    #print "Hello world"
    wb2 = openpyxl.load_workbook('1.xlsx')
#    for sheet in wb2.get_sheet_names():
#        print sheet
    
    for wsName in wb2.get_sheet_names():
        ws = wb2[wsName]
        cellIndex = 0
        toner = {}
        for cell in ws.columns[1]:
            try:
                #if datetime.datetime.strptime(cell.value, '%Y-%m-%d') > datetime.datetime.strptime('2015-05-01', '%Y-%m-%d'):
                
                if type(cell.value) == datetime.datetime:
                    if cell.value > datetime.datetime.strptime('2015-05-01', '%Y-%m-%d') and cell.value < datetime.datetime.strptime('2015-06-01', '%Y-%m-%d'):
                        toner["Date"] = datetime.datetime.strftime(cell.value,'%d-%m-%Y') 
                        toner["Status"] = ws.cell(row=cellIndex, column=2).value
                        toner["Recip"] = ws.cell(row=cellIndex, column=3).value
                        toner["Price"] = ws.cell(row=cellIndex-1, column=5).value
                        toner["SN"] = wsName
                        #print(type(toner["Price"]))
                        if toner["Status"].find(u'Вы') != -1:
                            #print(cell.value)
                            print(repr(toner).decode("unicode-escape"))
                         
                        #print(cell)
                cellIndex += 1       
            except Exception as ex:
                template = "An exception of type {0} occured. Arguments:\n{1!r}"
                message = template.format(type(ex).__name__, ex.args)
                print message     